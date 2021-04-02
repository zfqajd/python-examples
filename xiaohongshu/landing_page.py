#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# @Time     : 2021/4/2 11:06
# @Author   : Xo9
# @FileName : landing_page.py
# @Blog     : https://zfqajd.github.io

import os
import time
import httpx
import socket
import datetime
import SendMail
from openpyxl import load_workbook
from selenium import webdriver
from selenium.common.exceptions import *
from selenium.webdriver.common.by import By
from selenium.webdriver.firefox.options import Options
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.firefox.firefox_profile import FirefoxProfile

EXCEL_FILENAME = '/tmp/批量-数据.xlsx'


def login(proxy_ip, mail):
    LOGIN_URL = 'https://business.xiaohongshu.com/ares/login'
    LOGIN_EMAIL = ''
    LOGIN_PASS = ''
    GECKODRIVER_PATH = '/usr/local/bin/geckodriver'
    end_date = datetime.date.today()
    start_date = end_date - datetime.timedelta(days=7)
    year = start_date.strftime('%Y年')
    month = start_date.strftime('%-m月')
    day = start_date.strftime('%-d')

    if os.path.isfile(EXCEL_FILENAME):
        os.remove(EXCEL_FILENAME)

    # 配置代理
    # 参考：https://www.selenium.dev/documentation/zh-cn/webdriver/http_proxies/
    PROXY = "%s:8888" % proxy_ip
    webdriver.DesiredCapabilities.FIREFOX['proxy'] = {
        "httpProxy": PROXY,
        "ftpProxy": PROXY,
        "sslProxy": PROXY,
        "proxyType": "MANUAL",
    }

    # 启用无头模式
    # 参考：https://www.selenium.dev/documentation/zh-cn/driver_idiosyncrasies/driver_specific_capabilities/#使用firefoxoptions定义功能
    options = Options()
    options.headless = True

    # 自定义 User-Agent
    # 参数参考：http://kb.mozillazine.org/About:config_entries
    profile = FirefoxProfile()
    profile.set_preference("general.useragent.override", "Mozilla/5.0 (Macintosh; Intel Mac OS X 10.15; rv:87.0) Gecko/20100101 Firefox/87.0")
    # 变更下载目录
    profile.set_preference("browser.download.folderList", 2)
    profile.set_preference("browser.download.dir", "/tmp")
    profile.set_preference("browser.download.useDownloadDir", True)
    profile.set_preference("browser.download.manager.showWhenStarting", False)
    profile.set_preference("browser.helperApps.neverAsk.saveToDisk", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

    driver = webdriver.Firefox(executable_path=GECKODRIVER_PATH, firefox_profile=profile, options=options)
    driver.maximize_window()
    try:
        driver.get(LOGIN_URL)
        # 切换到 账号登录
        find_element(driver, "//div[normalize-space()='账号登录'][@class='css-1r2f04i']").click()
        # 输入邮箱
        email = find_element(driver, "//input[@placeholder='邮箱'][@class='dyn css-ley1ah css-14iyno2']")
        email.clear()
        email.send_keys(LOGIN_EMAIL)
        # 输入密码
        password = find_element(driver, "//input[@type='password']")
        password.clear()
        password.send_keys(LOGIN_PASS)
        # 点击登录
        find_element(driver, "//button[@class='dyn css-z31jkd css-ecaz6i']").click()

        # 关闭广告
        ad = find_element(driver, "//img[@class='remove']")
        if ad:
            ad.click()

        # 已读通知
        notice = find_element(driver, "//button[@class='btn-primary'][@normalize-space()='已读']")
        if notice:
            notice.click()

        # 切换到 推广中心
        find_element(driver, "//ul[@class='header-tab']/li[2]").click()
        time.sleep(3)
        # 选择 投放工具——落地页
        find_element(driver, "//div[@class='css-1ppmvr3'][normalize-space()='投放工具']").click()
        find_element(driver, "//div[@class='css-1ppmvr3'][normalize-space()='落地页']").click()
        time.sleep(3)
        # 全选
        find_element(driver, "//table[@class='table-large']/thead/tr/th[1]/div").click()
        # 点击 下载数据
        find_element(driver, "//button[@type='button'][@owl='download']").click()
        # 选择日期
        find_element(driver, "//input[@class='date-display']").click()

        if driver.find_element_by_xpath("//a[@class='yam-calendar-year-select']").text != year:
            find_element(driver, "//a[@class='yam-calendar-year-select']").click()
            select_year(driver, year)

        if driver.find_element_by_xpath("//a[@class='yam-calendar-month-select']").text != month:
            find_element(driver, "//a[@class='yam-calendar-month-select']").click()
            select_month(driver, month)

        days_table = driver.find_element_by_xpath("//table[@class='days']")
        days_tr = days_table.find_elements_by_tag_name("tr")
        for tr in days_tr:
            days_td = tr.find_elements_by_xpath("td[@class='']")
            for td in days_td:
                if td.text == day:
                    td.click()
                    td.click()
                    break

        time.sleep(3)
        find_element(driver, "//button[@type='button'][@owl='modal-confirm']").click()
        time.sleep(10)
        guixiaotu()
    except NoSuchElementException:
        print("没有找到元素！")
        mail.send_mail("小红书预约失败", "未找到元素！")
    finally:
        driver.quit()


def find_element(driver, selector):
    try:
        element = WebDriverWait(driver, 10).until(EC.element_to_be_clickable((By.XPATH, selector)))
        return element
    except TimeoutException:
        return False


def select_year(driver, year):
    year_elements = driver.find_elements_by_xpath("//a[@class='yam-calendar-year-panel-year']")
    if year_elements:
        for element in year_elements:
            if element.text == year:
                element.click()
                break
            else:
                find_element(driver, "//a[@class='yam-calendar-year-panel-prev']").click()
                select_year(driver, year)


def select_month(driver, month):
    month_elements = driver.find_elements_by_xpath("//a[@class='yam-calendar-month-panel-month']")
    if month_elements:
        for element in month_elements:
            if element.text == month:
                element.click()
                break


def guixiaotu():
    API_URL = ''
    wb = load_workbook(filename=EXCEL_FILENAME)
    ws = wb.active
    if ws.max_row > 1:
        with httpx.Client() as client:
            for row in range(2, ws.max_row + 1):
                mobile = ws['J%s' % row].value
                appt_time = ws['C%s' % row].value
                grade = ws['K%s' % row].value
                payload = {
                    'name': 'None',
                    'mobile': mobile,
                    'weixin': mobile,
                    'time': appt_time,
                    'grade': grade,
                    'searchWord': '',
                    'utmSource': 'xiaohongshu',
                    'utmMedium': 'lead_ads',
                    'utmCampaign': '5_free_course',
                    'originHost': 'xiaohongshu'
                }

                r = client.post(API_URL, params=payload)
                if r.status_code == httpx.codes.OK:
                    print('%s 在 %s 预约了体验课，小朋友读 %s' % (mobile, appt_time, grade))
    else:
        print('没有新用户预约体验课！')


def get_proxy():
    url = ''
    proxy_ipaddr = socket.getaddrinfo(url, 'http')
    if proxy_ipaddr[0][4][0][:3] == "27.":
        proxy_ip = proxy_ipaddr[0][4][0]
        return proxy_ip
    else:
        return False


if __name__ == '__main__':
    print("脚本执行时间：%s" % time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
    proxy_ip = get_proxy()
    mail = SendMail.SendMail()
    print("代理IP是：%s" % proxy_ip)
    if proxy_ip:
        login(proxy_ip, mail)
    else:
        print("获取代理IP失败！")
        mail.send_mail("小红书预约失败", "获取代理IP地址失败！")
